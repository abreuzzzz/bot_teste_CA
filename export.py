"""Export de lançamentos para XLSX (com fallback CSV)."""
import io, csv, re
from datetime import date, timedelta
from consulta_financeira import _buscar, _valor, _valor_pago

MESES = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

TIPO_MAP = {
    "receber": "RECEBER", "receita": "RECEBER",
    "pagar":   "PAGAR",   "despesa": "PAGAR",   "pago": "PAGAR",
}
STATUS_MAP = {
    "atrasado": "ATRASADO", "recebido": "RECEBIDO",
    "pago":     "RECEBIDO", "aberto":   "EM_ABERTO",
    "pendente": "EM_ABERTO",
}


def parse_periodo(texto: str) -> tuple[date, date]:
    """Aceita 'mai2026', '05/2026', '2026-05', 'mai/26'. Default: mês atual."""
    # Remove palavras de tipo/status para não atrapalhar parsing de data
    palavras_filtro = set(TIPO_MAP) | set(STATUS_MAP)
    t = " ".join(w for w in (texto or "").lower().split() if w not in palavras_filtro)
    t = t.strip().replace("/", "").replace("-", "").replace(" ", "")
    hoje = date.today()
    m = re.match(r"^([a-z]{3})(\d{2,4})$", t)
    if m:
        mes = MESES.get(m.group(1))
        ano = int(m.group(2))
        if ano < 100:
            ano += 2000
        if mes:
            ini = date(ano, mes, 1)
            fim = (date(ano + 1, 1, 1) - timedelta(days=1)) if mes == 12 \
                  else (date(ano, mes + 1, 1) - timedelta(days=1))
            return ini, fim
    m = re.match(r"^(\d{1,2})(\d{4})$", t)
    if m:
        mes, ano = int(m.group(1)), int(m.group(2))
        ini = date(ano, mes, 1)
        fim = (date(ano + 1, 1, 1) - timedelta(days=1)) if mes == 12 \
              else (date(ano, mes + 1, 1) - timedelta(days=1))
        return ini, fim
    m = re.match(r"^(\d{4})(\d{2})$", t)
    if m:
        ano, mes = int(m.group(1)), int(m.group(2))
        ini = date(ano, mes, 1)
        fim = (date(ano + 1, 1, 1) - timedelta(days=1)) if mes == 12 \
              else (date(ano, mes + 1, 1) - timedelta(days=1))
        return ini, fim
    ini = hoje.replace(day=1)
    fim = (ini.replace(year=ini.year + 1, month=1, day=1) - timedelta(days=1)) if ini.month == 12 \
          else (ini.replace(month=ini.month + 1, day=1) - timedelta(days=1))
    return ini, fim


def parse_export_args(texto: str) -> dict:
    """
    Extrai período, tipo (RECEBER/PAGAR/AMBOS) e status da string de export.
    Ex: "mai2026 pagar atrasado" -> {ini, fim, tipo="PAGAR", status=["ATRASADO"]}
    """
    palavras = (texto or "").lower().split()
    tipo     = "AMBOS"
    status: list[str] = []
    for p in palavras:
        if p in TIPO_MAP:
            tipo = TIPO_MAP[p]
        elif p in STATUS_MAP and STATUS_MAP[p] not in status:
            status.append(STATUS_MAP[p])
    ini, fim = parse_periodo(texto)
    return {"ini": ini, "fim": fim, "tipo": tipo, "status": status or None}


def _coletar(ini: date, fim: date,
             tipo: str = "AMBOS", status: list[str] | None = None) -> list[dict]:
    p: dict = {"data_vencimento_de": str(ini), "data_vencimento_ate": str(fim)}
    if status:
        p["status"] = status
    linhas = []
    fontes = []
    if tipo in ("RECEBER", "AMBOS"):
        fontes.append(("RECEBER", "contas-a-receber/buscar"))
    if tipo in ("PAGAR", "AMBOS"):
        fontes.append(("PAGAR", "contas-a-pagar/buscar"))
    for tipo_ep, ep in fontes:
        for i in _buscar(ep, p):
            rateio   = i.get("rateio") or []
            categoria = (rateio[0].get("categoria") or {}).get("nome", "") if rateio else ""
            linhas.append({
                "tipo":            tipo_ep,
                "descricao":       i.get("descricao", ""),
                "valor":           round(_valor(i), 2),
                "valor_pago":      round(_valor_pago(i), 2),
                "data_vencimento": i.get("data_vencimento", ""),
                "data_pagamento":  i.get("data_pagamento", "") or "",
                "status":          i.get("status", ""),
                "categoria":       categoria,
                "contato":         (i.get("contato") or {}).get("nome", "") if isinstance(i.get("contato"), dict) else "",
            })
    linhas.sort(key=lambda x: x["data_vencimento"])
    return linhas


def exportar_xlsx(ini: date, fim: date,
                  tipo: str = "AMBOS", status: list[str] | None = None) -> tuple[bytes, str]:
    linhas = _coletar(ini, fim, tipo=tipo, status=status)
    sufixo = f"_{tipo.lower()}" if tipo != "AMBOS" else ""
    nome = f"lancamentos_{ini.strftime('%Y%m')}{sufixo}.xlsx"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Lançamentos"
        headers = ["Tipo", "Descrição", "Valor", "Valor Pago", "Vencimento",
                   "Pagamento", "Status", "Categoria", "Contato"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
        for l in linhas:
            ws.append([l[k] for k in
                       ("tipo", "descricao", "valor", "valor_pago",
                        "data_vencimento", "data_pagamento", "status",
                        "categoria", "contato")])
        for col in ws.columns:
            largura = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(largura + 2, 40)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), nome
    except ImportError:
        return exportar_csv(ini, fim, tipo=tipo, status=status)


def exportar_csv(ini: date, fim: date,
                 tipo: str = "AMBOS", status: list[str] | None = None) -> tuple[bytes, str]:
    linhas = _coletar(ini, fim, tipo=tipo, status=status)
    sufixo = f"_{tipo.lower()}" if tipo != "AMBOS" else ""
    nome = f"lancamentos_{ini.strftime('%Y%m')}{sufixo}.csv"
    buf = io.StringIO()
    if linhas:
        w = csv.DictWriter(buf, fieldnames=list(linhas[0].keys()), delimiter=";")
        w.writeheader()
        w.writerows(linhas)
    return buf.getvalue().encode("utf-8-sig"), nome
